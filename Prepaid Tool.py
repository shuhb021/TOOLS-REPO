import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import threading
import requests
import sys
import os
import subprocess
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill

"""
Prepaid Expense Workpaper Automation – GUI Module
All UI layout, widgets, and visual logic live here.
"""


CURRENT_VERSION = "v1.1.0"
GITHUB_REPO = "shuhb021/Prepaid-Tool"


# ─────────────────── Color Palette ───────────────────
DARK_THEME    = "#3A1C5E"   # Header
MID_THEME     = "#6B37A8"   # Button hover
ACCENT_THEME  = "#512882"   # Brand Purple
LIGHT_THEME   = "#EDE5F5"   # Nav active hover & Progress BG
BANNER_BG     = "#F0E6FF"
BANNER_TEXT   = "#512882"
MAIN_BG       = "#F4F5F7"   # Light Gray/White Main Background
CARD_BG       = "#FFFFFF"   # Pure White Cards
STAT_HEADER   = "#512882"
BORDER_CLR    = "#D4C8E8"
MUTED_TEXT    = "#6B7280"
GREEN_OK      = "#16A34A"
STATUS_COLOR  = "#856CA9"
TEXT_DARK     = "#111827"   # Dark text for light background
SEPARATOR_CLR = "#E5E7EB"

NAV_ITEMS = []


class PrepaidApp(ctk.CTk):
    """Modern Prepaid Expense Workpaper Automation tool.

    Parameters
    ----------
    process_fn : callable
        The backend processing function with signature::

            process_fn(basefile, client_name, fy_start, fy_end,
                       method, sample_n, progress_callback)
            -> (total_rows, sample_total, addition_count,
                addition_sum, addition_pct, output_path)

        ``progress_callback(value: float, text: str)`` is called by the
        backend to report progress (0.0 – 1.0).
    """

    def __init__(self, process_fn=None):
        super().__init__()

        self.process_fn = process_fn

        self.title("Prepaid Expense Workpaper Automation")
        self.geometry("1100x720")
        self.minsize(1050, 680)
        self.configure(fg_color=MAIN_BG)

        ctk.set_appearance_mode("light")

        # Layout: 2 columns (sidebar | main), 2 rows (header | body)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_header()
        self._build_sidebar()
        self._build_main()

        # Check for updates in background
        threading.Thread(target=self._check_updates_bg, daemon=True).start()

    # ─────────────── Auto-Updater ───────────────
    def _check_updates_bg(self):
        try:
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            response = requests.get(url, timeout=5)
            response.raise_for_status()
            data = response.json()
            latest_version = data.get("tag_name")
            
            if latest_version and latest_version != CURRENT_VERSION:
                self.after(0, self._prompt_update, latest_version, data)
        except requests.exceptions.HTTPError as e:
            if e.response.status_code != 404:
                print("Update check failed:", e)
        except requests.exceptions.RequestException:
            pass # Ignore network errors silently
        except Exception as e:
            print("Update check failed:", e)

    def _prompt_update(self, latest_version, data):
        ans = messagebox.askyesno(
            "Update Available",
            f"A new version ({latest_version}) is available. You are currently on {CURRENT_VERSION}.\n\nDo you want to update now?"
        )
        if ans:
            assets = data.get("assets", [])
            exe_url = None
            for asset in assets:
                if asset["name"].endswith(".exe"):
                    exe_url = asset["browser_download_url"]
                    break
            
            if exe_url:
                threading.Thread(target=self._download_and_update, args=(exe_url,), daemon=True).start()
            else:
                messagebox.showerror("Update Error", "No executable found in the latest release.")

    def _download_and_update(self, exe_url):
        self.after(0, lambda: self.btn_generate.configure(state="disabled", text="Downloading Update..."))
        try:
            response = requests.get(exe_url, stream=True)
            response.raise_for_status()
            
            new_exe_path = "Prepaid Tool_new.exe"
            with open(new_exe_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            bat_path = "update.bat"
            current_exe = sys.executable
            if not getattr(sys, 'frozen', False):
                self.after(0, lambda: messagebox.showinfo("Update", "Update downloaded, but you are not running the compiled .exe"))
                self.after(0, lambda: self.btn_generate.configure(state="normal", text="  \U0001F4CB  Generate Workpaper"))
                return
                
            exe_name = os.path.basename(current_exe)
            
            bat_content = f"""@echo off
echo Updating Prepaid Tool... Please wait.
timeout /t 2 /nobreak > NUL
move /y "{new_exe_path}" "{exe_name}"
start "" "{exe_name}"
del "%~f0"
"""
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write(bat_content)
                
            subprocess.Popen([bat_path], shell=True)
            self.after(0, self.destroy)
            
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Update Failed", f"Failed to download update: {e}"))
            self.after(0, lambda: self.btn_generate.configure(state="normal", text="  \U0001F4CB  Generate Workpaper"))

    # ─────────────── Header ───────────────
    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color=DARK_THEME, corner_radius=0, height=45)
        header.grid(row=0, column=0, columnspan=2, sticky="ew")
        header.grid_propagate(False)

        ctk.CTkLabel(
            header,
            text="   Prepaid Expense Workpaper Automation",
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            text_color="white",
        ).pack(side="left", padx=10, pady=8)

    # ─────────────── Sidebar ───────────────
    def _build_sidebar(self):
        # Outer wrapper gives a purple right-edge accent line
        sidebar_wrapper = ctk.CTkFrame(self, fg_color=ACCENT_THEME, corner_radius=0, width=224)
        sidebar_wrapper.grid(row=1, column=0, sticky="nsew")
        sidebar_wrapper.grid_propagate(False)

        # Inner white panel (3px right margin exposes the purple accent)
        sidebar = ctk.CTkFrame(sidebar_wrapper, fg_color=CARD_BG, corner_radius=0)
        sidebar.pack(side="left", fill="both", expand=True, padx=(0, 3))

        ctk.CTkLabel(
            sidebar,
            text="Walker Chandiok\n& Co LLP",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color=TEXT_DARK,
            justify="left",
        ).pack(anchor="w", padx=22, pady=(22, 28))

        self.nav_buttons = []
        for i, (icon, label) in enumerate(NAV_ITEMS):
            is_active = i == 0
            btn = ctk.CTkButton(
                sidebar,
                text=f" {icon}  {label}",
                font=ctk.CTkFont(family="Segoe UI", size=14),
                fg_color=ACCENT_THEME if is_active else "transparent",
                hover_color=LIGHT_THEME,
                text_color="white" if is_active else ACCENT_THEME,
                anchor="w",
                height=42,
                corner_radius=8,
                command=lambda idx=i: self._nav_click(idx),
            )
            btn.pack(fill="x", padx=14, pady=3)
            self.nav_buttons.append(btn)

        ctk.CTkLabel(
            sidebar,
            text="Walker Chandiok\n& Co LLP",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=MUTED_TEXT,
            justify="left",
        ).pack(side="bottom", anchor="w", padx=22, pady=22)

    def _nav_click(self, idx):
        for i, btn in enumerate(self.nav_buttons):
            if i == idx:
                btn.configure(fg_color=ACCENT_THEME, text_color="white")
            else:
                btn.configure(fg_color="transparent", text_color=ACCENT_THEME)

    # ─────────────── Main Content ───────────────
    def _build_main(self):
        main = ctk.CTkFrame(self, fg_color=MAIN_BG, corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew")
        main.grid_columnconfigure(0, weight=3)
        main.grid_columnconfigure(1, weight=2)
        main.grid_rowconfigure(1, weight=1)

        # ── Top bar: Client + Status ──
        top_bar = ctk.CTkFrame(main, fg_color=MAIN_BG, height=40)
        top_bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=25, pady=(15, 5))

        self.client_header = ctk.CTkLabel(
            top_bar,
            text="Client: \u2014",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color=TEXT_DARK,
        )
        self.client_header.pack(side="left")

        self.status_dot = ctk.CTkLabel(
            top_bar,
            text="\u25CF  Not Started",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=STATUS_COLOR,
        )
        self.status_dot.pack(side="right")

        # ── Left – scrollable form ──
        left = ctk.CTkScrollableFrame(main, fg_color=MAIN_BG, corner_radius=0)
        left.grid(row=1, column=0, sticky="nsew", padx=(25, 10), pady=5)

        # Banner
        banner = ctk.CTkFrame(left, fg_color=BANNER_BG, corner_radius=10, height=50)
        banner.pack(fill="x", pady=(0, 18))
        ctk.CTkLabel(
            banner,
            text="  \U0001F4CB  Automate prepaid expense calculations and generate workpapers efficiently.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=BANNER_TEXT,
            wraplength=500,
        ).pack(padx=15, pady=12, anchor="w")

        # ── Select Input File ──
        self._section_label(left, "Select Input File")
        file_row = ctk.CTkFrame(left, fg_color="transparent")
        file_row.pack(fill="x", pady=(0, 12))

        self.entry_file = ctk.CTkEntry(
            file_row, height=38, corner_radius=8,
            border_color=BORDER_CLR, fg_color=CARD_BG,
            placeholder_text="Select an Excel file...",
        )
        self.entry_file.pack(side="left", fill="x", expand=True, padx=(0, 8))

        ctk.CTkButton(
            file_row, text="Browse", width=90, height=38,
            fg_color=ACCENT_THEME, hover_color=MID_THEME,
            corner_radius=8, command=self._browse_file,
        ).pack(side="right")

        # ── Client Name ──
        self._section_label(left, "Client Name")
        self.entry_client = ctk.CTkEntry(
            left, height=38, corner_radius=8,
            border_color=BORDER_CLR, fg_color=CARD_BG,
            placeholder_text="Enter client name",
        )
        self.entry_client.pack(fill="x", pady=(0, 12))
        self.entry_client.bind("<FocusOut>", self._on_client_entered)

        # ── FY Start Date ──
        self._section_label(left, "FY Start Date (DD-MM-YYYY)")
        self.entry_fy_start = ctk.CTkEntry(
            left, height=38, corner_radius=8,
            border_color=BORDER_CLR, fg_color=CARD_BG,
            placeholder_text="DD-MM-YYYY",
        )
        self.entry_fy_start.pack(fill="x", pady=(0, 12))
        self.entry_fy_start.bind("<FocusOut>", self._on_start_entered)

        # ── FY End Date ──
        self._section_label(left, "FY End Date (DD-MM-YYYY)")
        self.entry_fy_end = ctk.CTkEntry(
            left, height=38, corner_radius=8,
            border_color=BORDER_CLR, fg_color=CARD_BG,
            placeholder_text="DD-MM-YYYY",
        )
        self.entry_fy_end.pack(fill="x", pady=(0, 15))
        self.entry_fy_end.bind("<FocusOut>", self._on_end_entered)

        # ── Sampling Card ──
        sampling_card = ctk.CTkFrame(
            left, fg_color=CARD_BG, corner_radius=10,
            border_color=BORDER_CLR, border_width=1,
        )
        sampling_card.pack(fill="x", pady=(0, 15))

        self.sampling_method = ctk.StringVar(value="top")

        ctk.CTkRadioButton(
            sampling_card, text="Top-N Sampling",
            variable=self.sampling_method, value="top",
            font=ctk.CTkFont(size=13),
            fg_color=ACCENT_THEME, hover_color=MID_THEME,
            border_color=BORDER_CLR,
        ).pack(anchor="w", padx=20, pady=(15, 5))

        ctk.CTkFrame(sampling_card, fg_color=BORDER_CLR, height=1).pack(
            fill="x", padx=20, pady=5
        )

        ctk.CTkRadioButton(
            sampling_card, text="Random Sampling",
            variable=self.sampling_method, value="random",
            font=ctk.CTkFont(size=13),
            fg_color=ACCENT_THEME, hover_color=MID_THEME,
            border_color=BORDER_CLR,
        ).pack(anchor="w", padx=20, pady=5)

        ctk.CTkLabel(
            sampling_card, text="Enter the no of values to be picked:",
            font=ctk.CTkFont(size=12), text_color=TEXT_DARK,
        ).pack(anchor="w", padx=20, pady=(10, 4))

        self.entry_top_n = ctk.CTkEntry(
            sampling_card, height=35, width=120, corner_radius=8,
            border_color=BORDER_CLR, fg_color=CARD_BG,
        )
        self.entry_top_n.pack(anchor="w", padx=20, pady=(0, 15))

        # ── Generate Button ──
        self.btn_generate = ctk.CTkButton(
            left,
            text="  \U0001F4CB  Generate Workpaper",
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=ACCENT_THEME, hover_color=MID_THEME,
            height=45, corner_radius=10,
            command=self._run_process,
        )
        self.btn_generate.pack(anchor="w", pady=(5, 20))

        # ── Right – Statistics Panel ──
        right = ctk.CTkFrame(
            main, fg_color=CARD_BG, corner_radius=12,
            border_color=BORDER_CLR, border_width=1,
        )
        right.grid(row=1, column=1, sticky="nsew", padx=(10, 25), pady=5)

        ctk.CTkLabel(
            right, text="Statistics",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=STAT_HEADER,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        stat_items = [
            ("total_rows",     "Total Rows Processed:"),
            ("audit_period",   "Audit Period:"),
            ("addition_count", "No of Total Addition in the\nCurrent Year:"),
            ("addition_sum",   "Amount of Total Addition in the\nCurrent Year:"),
            ("sample_total",   "Sample Amount Total:"),
            ("sample_pct",     "Sample Amount % of Total:"),
            ("status",         "Status:"),
        ]

        self.stat_labels = {}
        for key, label_text in stat_items:
            row_f = ctk.CTkFrame(right, fg_color="transparent")
            row_f.pack(fill="x", padx=20, pady=6)

            ctk.CTkLabel(
                row_f, text=label_text,
                font=ctk.CTkFont(size=12), text_color="#4A4A5A",
                justify="left", wraplength=180,
            ).pack(side="left", anchor="nw")

            val_color = STATUS_COLOR if key == "status" else TEXT_DARK
            val_text  = "Not Started" if key == "status" else "-"
            lbl = ctk.CTkLabel(
                row_f, text=val_text,
                font=ctk.CTkFont(
                    size=12, weight="bold" if key == "status" else "normal"
                ),
                text_color=val_color,
            )
            lbl.pack(side="right", anchor="ne")
            self.stat_labels[key] = lbl

            if key != "status":
                ctk.CTkFrame(right, fg_color=SEPARATOR_CLR, height=1).pack(
                    fill="x", padx=20
                )

        # ── Progress bar (bottom) ──
        self.progress = ctk.CTkProgressBar(
            main, fg_color=LIGHT_THEME,
            progress_color=ACCENT_THEME,
            height=6, corner_radius=3,
        )
        self.progress.grid(
            row=2, column=0, columnspan=2, sticky="ew", padx=25, pady=(5, 5)
        )
        self.progress.set(0)

        self.status_label = ctk.CTkLabel(
            main, text="Waiting for input...",
            font=ctk.CTkFont(size=11), text_color=MUTED_TEXT,
        )
        self.status_label.grid(
            row=3, column=0, columnspan=2, sticky="w", padx=25, pady=(0, 10)
        )

    # ─────────────── UI Helpers ───────────────
    @staticmethod
    def _section_label(parent, text):
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=TEXT_DARK,
        ).pack(anchor="w", pady=(0, 4))

    def _browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.entry_file.delete(0, "end")
            self.entry_file.insert(0, path)
            self._update_progress(0.20, "Base file selected")

    def _on_client_entered(self, _event=None):
        name = self.entry_client.get()
        if name:
            self.client_header.configure(text=f"Client: {name}")
            self._update_progress(0.40, "Client name captured")

    def _on_start_entered(self, _event=None):
        if self.entry_fy_start.get():
            self._update_progress(0.60, "FY start date captured")

    def _on_end_entered(self, _event=None):
        if self.entry_fy_end.get():
            self._update_progress(0.80, "Audit period completed")

    def _update_progress(self, value, text):
        current = self.progress.get()
        if value > current or value == 0:
            self.progress.set(value)
            self.status_label.configure(text=text)
            self.update_idletasks()

    def _update_status_dot(self, text, color=STATUS_COLOR):
        self.status_dot.configure(text=f"\u25CF  {text}", text_color=color)

    def _show_statistics(self, total_rows, sample_total, addition_count,
                         addition_sum, addition_pct):
        self.stat_labels["total_rows"].configure(text=str(total_rows))
        fy_s = self.entry_fy_start.get()
        fy_e = self.entry_fy_end.get()
        self.stat_labels["audit_period"].configure(text=f"{fy_s} to {fy_e}")
        self.stat_labels["addition_count"].configure(
            text=str(round(addition_count, 2))
        )
        self.stat_labels["addition_sum"].configure(
            text=str(round(addition_sum, 2))
        )
        self.stat_labels["sample_total"].configure(
            text=str(round(sample_total, 2))
        )
        self.stat_labels["sample_pct"].configure(
            text=f"{round(addition_pct, 2)}%"
        )
        self.stat_labels["status"].configure(
            text="Completed \u2705", text_color=GREEN_OK
        )

    def _processing_complete(self, total_rows, sample_total, addition_count,
                             addition_sum, addition_pct, output_path):
        self._update_progress(1.0, "Processing completed \u2705")
        self._show_statistics(
            total_rows, sample_total, addition_count, addition_sum, addition_pct
        )
        self._update_status_dot("Completed", GREEN_OK)
        self.btn_generate.configure(
            state="normal", text="  \U0001F4CB  Generate Workpaper"
        )
        messagebox.showinfo(
            "Success",
            f"\u2705 Workpaper Generated!\nSaved at:\n{output_path}",
        )

    def _reset_ui(self):
        self.btn_generate.configure(
            state="normal", text="  \U0001F4CB  Generate Workpaper"
        )
        self._update_progress(0, "Waiting for input...")
        self._update_status_dot("Not Started")

    # ─────────────── Thread Launcher ───────────────
    def _run_process(self):
        if self.process_fn is None:
            messagebox.showerror("Error", "No processing function configured.")
            return

        basefile     = self.entry_file.get()
        client_name  = self.entry_client.get()
        fy_start     = self.entry_fy_start.get()
        fy_end       = self.entry_fy_end.get()
        method       = self.sampling_method.get()
        sample_n_str = self.entry_top_n.get()

        # ── Quick validation (UI-side) ──
        if not all([basefile, client_name, fy_start, fy_end]):
            messagebox.showerror("Error", "Please fill all fields")
            return
        if not sample_n_str:
            messagebox.showerror("Error", "Please enter the values for Sampling")
            return
        try:
            sample_n = int(sample_n_str)
        except ValueError:
            messagebox.showerror("Error", "Sample size must be an integer")
            return

        self.btn_generate.configure(state="disabled", text="  \u23F3  Processing...")
        self._update_status_dot("Processing...", "#F59E0B")

        def _thread_target():
            def progress_cb(value, text):
                self.after(0, self._update_progress, value, text)

            try:
                result = self.process_fn(
                    basefile, client_name, fy_start, fy_end,
                    method, sample_n, progress_cb,
                )
                self.after(0, self._processing_complete, *result)
            except Exception as e:
                self.after(
                    0, messagebox.showerror, "Error",
                    f"An error occurred:\n{str(e)}",
                )
                self.after(0, self._reset_ui)

        thread = threading.Thread(target=_thread_target, daemon=True)
        thread.start()

"""
Prepaid Expense Workpaper Automation – Business Logic & Entry Point
"""




# ──────────────────────────────────────────────────────────
#  Core processing function (runs in a background thread)
# ──────────────────────────────────────────────────────────
def process_workpaper(basefile, client_name, fy_start, fy_end,
                      method, sample_n, progress):
    """Generate the prepaid-expense workpaper.

    Parameters
    ----------
    basefile     : str   – path to the input .xlsx file
    client_name  : str   – client / firm name
    fy_start     : str   – FY start date  (DD-MM-YYYY)
    fy_end       : str   – FY end date    (DD-MM-YYYY)
    method       : str   – "top" or "random"
    sample_n     : int   – number of samples to pick
    progress     : callable(float, str) – callback to report progress

    Returns
    -------
    tuple : (total_rows, sample_total, addition_count,
             addition_sum, addition_pct, output_path)
    """

    output_path = basefile.replace(".xlsx", "_output.xlsx")

    fy_start_date = datetime.strptime(fy_start, "%d-%m-%Y")
    fy_end_date   = datetime.strptime(fy_end,   "%d-%m-%Y")
    start_year    = fy_start_date.year
    end_year      = fy_end_date.year

    # ── Read Template ──
    progress(0.10, "Reading template...")
    template_df = pd.read_excel(basefile, sheet_name="Template", header=1)
    template_df = template_df.iloc[1:]
    template_df.columns = (
        template_df.columns.astype(str)
        .str.replace("\n", " ")
        .str.strip()
    )
    if "Expense" in template_df.columns:
        template_df = template_df.rename(columns={"Expense": "Particulars"})
    template_df = template_df.dropna(how="all").reset_index(drop=True)

    # Drop pre-numbered template rows that carry only a serial number ("S.no")
    # and no real data.  These otherwise get written out and leave a long
    # empty tail below the TOTAL row.
    meaningful_cols = [
        c for c in template_df.columns
        if str(c).strip().lower() != "s.no" and not str(c).startswith("Unnamed")
    ]
    if meaningful_cols:
        template_df = (
            template_df.dropna(how="all", subset=meaningful_cols)
            .reset_index(drop=True)
        )

    # ── Load Workbook ──
    progress(0.20, "Loading workbook...")
    wb  = load_workbook(basefile)
    ws  = wb["Prepaid expenses"]
    ws2 = wb["Lead"]

    # ── Header cells ──
    ws["C2"]  = client_name
    ws["C3"]  = f"{fy_start} to {fy_end}"
    ws["Q38"] = fy_start_date
    ws["R38"] = fy_end_date
    ws2["C2"] = client_name
    ws2["C3"] = f"{fy_start} to {fy_end}"
    
    # Previous FY date and Current FY date for Leads sheet
    ws2["C20"] = fy_end_date 
    ws2["D20"] = fy_end_date - relativedelta(years=1)
    
    def update_dynamic_dates(value):
        if value is None:
            return value
        if not isinstance(value, str):
            return value

        text = value
        start_dm = fy_start_date.strftime("%d %B").lstrip("0")
        end_dm = fy_end_date.strftime("%d %B").lstrip("0")
        start_month = fy_start_date.strftime("%B")
        end_month = fy_end_date.strftime("%B")
        months = r"(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
        
        # Replace Full Date + Year
        text = re.sub(rf"0?1(?:st)?[\s\-]+{months}[\s\-]+(20XX|\d{{4}})", f"{start_dm} {start_year}", text, flags=re.IGNORECASE)
        text = re.sub(rf"(?:28|29|30|31)(?:st|nd|rd|th)?[\s\-]+{months}[\s\-]+(20XX|\d{{4}})", f"{end_dm} {end_year}", text, flags=re.IGNORECASE)
        
        # Replace Date + Month (without year)
        text = re.sub(rf"0?1(?:st)?[\s\-]+{months}", start_dm, text, flags=re.IGNORECASE)
        text = re.sub(rf"(?:28|29|30|31)(?:st|nd|rd|th)?[\s\-]+{months}", end_dm, text, flags=re.IGNORECASE)
        
        # Fallback year replacement for standalone 20XX or 202X
        if end_month in text:
            text = text.replace("20XX", str(end_year))
        elif start_month in text:
            text = text.replace("20XX", str(start_year))
        else:
            text = text.replace("20XX", str(end_year))
            
        return text

    # ── Map columns ──
    progress(0.30, "Mapping columns...")
    output_headers = [cell.value for cell in ws[40]]
    output_headers = [
        str(c).replace("\n", " ").strip() if c else None
        for c in output_headers
    ]
    output_col_map = {
        col: idx + 1 for idx, col in enumerate(output_headers) if col
    }
    special_col_mapping = {
        "Document No/ Invoice No.": "Document No",
        "Party Name": "Vendor Name",
    }

    # ── Prepare data ──
    progress(0.40, "Preparing data...")
    start_row = 42

    write_ops = []
    for i, row in template_df.iterrows():
        excel_row = start_row + i
        ssd = pd.to_datetime(row.get("Service start date"), errors="coerce")
        valid_addition = pd.notna(ssd) and ssd >= fy_start_date

        for col_name in template_df.columns:
            if col_name not in output_col_map:
                continue
            col_idx = output_col_map[col_name]

            if col_name == "Addition during the year":
                value = row.get("Amount", None) if valid_addition else None
            elif col_name in special_col_mapping:
                value = row.get(special_col_mapping[col_name])
            else:
                value = row.get(col_name)

            value = update_dynamic_dates(value)
            template_df.at[i, col_name] = value
            write_ops.append((excel_row, col_idx, value))

    # ── Write data ──
    progress(0.50, "Writing data to Excel...")
    for r, c, v in write_ops:
        ws.cell(row=r, column=c, value=v)

    # ── Determine the last real data row dynamically from the
    #    Service Start / End Date columns (TOTAL goes immediately after it) ──
    start_col_idx = next(
        (i for c, i in output_col_map.items() if "service start date" in str(c).lower()), 17
    )
    end_col_idx = next(
        (i for c, i in output_col_map.items() if "service end date" in str(c).lower()), 18
    )

    def is_empty(val):
        """A cell is empty when it has no real value: None, NaN/NaT/NA,
        an empty string, or whitespace only.  Formatting is ignored."""
        if val is None:
            return True
        try:
            if pd.isna(val):          # float NaN, NaT, pandas NA
                return True
        except (TypeError, ValueError):
            pass
        return isinstance(val, str) and val.strip() == ""

    last_start_row = last_end_row = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        if not is_empty(ws.cell(row=r, column=start_col_idx).value):
            last_start_row = max(last_start_row, r)
        if not is_empty(ws.cell(row=r, column=end_col_idx).value):
            last_end_row = max(last_end_row, r)

    last_data_row = max(last_start_row, last_end_row)
    last_row = last_data_row if last_data_row >= start_row else start_row - 1

    print(f"Last Service Start Date Row = {last_start_row}")
    print(f"Last Service End Date Row = {last_end_row}")
    print(f"Last Data Row = {last_row}")

    # ── Formulas ──
    progress(0.60, "Applying formulas...")
    for row in range(start_row, last_row + 1):
        ws[f"N{row}"]  = f"=I{row}-M{row}"
        ws[f"O{row}"]  = f"=R{row}-Q{row}+1"
        ws[f"P{row}"]  = (
            f"=IF(R{row}<$Q$38,0,"
            f"IF(Q{row}<$Q$38,R{row}-$Q$38+1,O{row}))"
        )
        ws[f"S{row}"]  = (
            f"=IF(Q{row}>$Q$38,"
            f"IF(R{row}>$R$38,$R$38-Q{row}+1,R{row}-Q{row}+1),"
            f"IF(R{row}>$R$38,$R$38-$Q$38+1,R{row}-$Q$38+1))"
        )
        ws[f"T{row}"]  = f"=P{row}-S{row}"
        ws[f"U{row}"]  = f"=I{row}*(S{row}/O{row})"
        ws[f"V{row}"]  = f"=ROUND((T{row}/O{row})*I{row},2)"
        ws[f"W{row}"]  = f"=L{row}-U{row}"
        ws[f"X{row}"]  = f"=N{row}-V{row}"
        ws[f"AA{row}"] = f"=IF(T{row}>365,(V{row}*(365/T{row})),V{row})"
        ws[f"AB{row}"] = f"=V{row}-AA{row}"

    # ── Formatting ──
    progress(0.70, "Applying formatting...")
    col_styles = {}
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=42, column=col)
        if src.has_style:
            col_styles[col] = src._style
    if col_styles:
        for row in range(start_row, last_row + 2):
            for col, style in col_styles.items():
                ws.cell(row=row, column=col)._style = style

    # ── Update dynamic dates across all sheets ──
    progress(0.80, "Updating header dates...")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            if not any(c.value is not None for c in row):
                continue
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value
                    cell.value = update_dynamic_dates(val)

    # ── Totals ──
    # TOTAL row sits immediately after the last real data row.
    total_row = last_row + 1
    print(f"Total Row = {total_row}")

    coord27_s = ws.cell(row=start_row, column=27).coordinate
    coord27_e = ws.cell(row=last_row,  column=27).coordinate
    coord28_s = ws.cell(row=start_row, column=28).coordinate
    coord28_e = ws.cell(row=last_row,  column=28).coordinate

    # Wipe any leftover template formulas/values sitting on the TOTAL row so it
    # shows only the "Total" label and the Current / Non-Current sums.
    for col in range(1, ws.max_column + 1):
        ws.cell(row=total_row, column=col).value = None

    ws["E32"] = f"=SUM({coord27_s}:{coord27_e})"
    ws["E33"] = f"=SUM({coord28_s}:{coord28_e})"
    ws.cell(row=total_row, column=1).value = "Total"           # column A label
    ws.cell(row=total_row, column=27).value = f"=SUM({coord27_s}:{coord27_e})"
    ws.cell(row=total_row, column=28).value = f"=SUM({coord28_s}:{coord28_e})"

    # ── Sampling ──
    progress(0.90, "Performing sampling...")
    data  = ws.iter_rows(min_row=start_row, max_row=last_row, values_only=True)
    df_ws = pd.DataFrame(data, columns=output_headers)

    df_ws["_sampling_amount_"] = None
    ssd_col = pd.to_datetime(df_ws["Service start date"], errors="coerce")
    mask = ssd_col >= fy_start_date
    df_ws.loc[mask, "_sampling_amount_"] = df_ws.loc[mask, "Amount"]

    if method == "top":
        sampled_indices = (
            df_ws[df_ws["_sampling_amount_"].notna()]
            .sort_values(by="_sampling_amount_", ascending=False)
            .head(sample_n)
            .index
        )
    elif method == "random":
        eligible_df = df_ws[df_ws["_sampling_amount_"].notna()]
        if sample_n > len(eligible_df):
            raise ValueError("Sample size exceeds eligible records")
        sampled_indices = eligible_df.sample(n=sample_n, random_state=42).index
    else:
        sampled_indices = []

    doc_col_idx   = output_col_map.get("Document No/ Invoice No.")
    party_col_idx = output_col_map.get("Party Name")
    doc_src_idx   = (df_ws.columns.get_loc("Document No")
                     if "Document No" in df_ws.columns else None)
    party_src_idx = (df_ws.columns.get_loc("Vendor Name")
                     if "Vendor Name" in df_ws.columns else None)

    # Clear all first
    for i in range(len(df_ws)):
        er = start_row + i
        if doc_col_idx:
            ws.cell(row=er, column=doc_col_idx, value=None)
        if party_col_idx:
            ws.cell(row=er, column=party_col_idx, value=None)

    # Write only sampled rows
    for i in sampled_indices:
        er = start_row + i
        if doc_col_idx and doc_src_idx is not None:
            ws.cell(row=er, column=doc_col_idx,
                    value=df_ws.iloc[i, doc_src_idx])
        if party_col_idx and party_src_idx is not None:
            ws.cell(row=er, column=party_col_idx,
                    value=df_ws.iloc[i, party_src_idx])

    # Statistics
    sample_total = df_ws.loc[sampled_indices, "_sampling_amount_"].sum()
    add_series = pd.to_numeric(
        template_df["Addition during the year"].dropna(), errors="coerce"
    ).dropna()
    addition_sum   = add_series.sum()
    addition_count = len(add_series)
    addition_pct   = (sample_total / addition_sum * 100) if addition_sum else 0

    # Total-row formatting: bold text, NO grey background fill.
    bottom_row = total_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=bottom_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type=None)   # remove the grey background
        if col >= 2:
            cell.number_format = "#,##0"

    # ── Force the Service Start / End Date columns to a real date format
    #    (match the Template's "d-mmm-yy") so they don't show as serial numbers ──
    DATE_FMT = "d-mmm-yy"
    for r in range(start_row, last_row + 1):
        ws.cell(row=r, column=start_col_idx).number_format = DATE_FMT
        ws.cell(row=r, column=end_col_idx).number_format = DATE_FMT

    # ── End the sheet right after the TOTAL row: remove every trailing
    #    template row below it so nothing prints after the totals ──
    if ws.max_row > total_row:
        ws.delete_rows(total_row + 1, ws.max_row - total_row)

    # Remove Template sheet & unhide all
    if "Template" in wb.sheetnames:
        del wb["Template"]
    for sheet in wb.worksheets:
        sheet.sheet_state = "visible"

    # Save
    progress(0.95, "Saving output file...")
    wb.save(output_path)

    return (
        len(template_df),
        sample_total,
        addition_count,
        addition_sum,
        addition_pct,
        output_path,
    )


# ──────────────────────────────────────────────────────────
#  Entry Point
# ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = PrepaidApp(process_fn=process_workpaper)
    app.mainloop()